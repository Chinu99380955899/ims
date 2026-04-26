"""Invoice upload, listing, detail, and export endpoints."""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date
from math import ceil
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, File, Query, UploadFile, status
from fastapi.responses import StreamingResponse

from app.api.deps import CurrentUser, DBSession
from app.core.config import settings
from app.core.logging import get_logger
from app.db.models.invoice import InvoiceStatus
from app.schemas.common import Page
from app.schemas.invoice import (
    DashboardStats,
    InvoiceDetail,
    InvoiceFilters,
    InvoiceRead,
    InvoiceUploadResponse,
)
from app.services.invoice_service import InvoiceService
from app.services.storage_service import get_storage
from app.utils.exceptions import (
    FileTooLargeError,
    UnsupportedFileTypeError,
)
from app.utils.hashing import sha256_of_bytes

router = APIRouter(prefix="/invoices", tags=["invoices"])
log = get_logger(__name__)


def _validate_upload(file: UploadFile, size: int) -> None:
    # 1. Size Validation
    if size > settings.max_upload_bytes:
        raise FileTooLargeError(
            f"File exceeds max size of {settings.max_upload_size_mb} MB",
            details={"size": size, "max": settings.max_upload_bytes},
        )

    # 2. Strict MIME Type Checking (SECURITY + PDF SUPPORT)
    allowed_mimes = {"application/pdf", "image/png", "image/jpeg", "image/tiff"}
    if file.content_type not in allowed_mimes:
        raise UnsupportedFileTypeError(
            f"MIME type '{file.content_type}' is not supported.",
            details={"allowed_mimes": list(allowed_mimes)},
        )

    # 3. Extension Validation
    filename = (file.filename or "").lower()
    ext = filename.rsplit(".", 1)[-1] if "." in filename else ""
    guaranteed_exts = {"pdf", "png", "jpg", "jpeg", "tiff", "tif"}
    allowed_exts = set(settings.allowed_extensions_list).union(guaranteed_exts)
    
    if ext not in allowed_exts:
        raise UnsupportedFileTypeError(
            f"Extension '.{ext}' is not allowed",
            details={"allowed": list(allowed_exts)},
        )


@router.post(
    "/upload",
    response_model=InvoiceUploadResponse,
    status_code=status.HTTP_202_ACCEPTED,
)
async def upload_invoice(
    db: DBSession,
    user: CurrentUser,
    file: UploadFile = File(..., description="Invoice file (PDF/PNG/JPEG/TIFF)"),
) -> InvoiceUploadResponse:
    content = await file.read()
    _validate_upload(file, len(content))

    file_hash = sha256_of_bytes(content)
    storage = get_storage()
    key = f"invoices/{file_hash[:2]}/{file_hash}_{file.filename}"
    storage_path = storage.save(key, content)

    service = InvoiceService(db)
    invoice = service.create(
        original_filename=file.filename or "unknown",
        storage_path=storage_path,
        file_hash=file_hash,
        file_size_bytes=len(content),
        mime_type=file.content_type or "application/octet-stream",
        uploaded_by_id=user.id,
    )

    from app.workers.tasks import process_invoice_task
    process_invoice_task.apply_async(
        args=[str(invoice.id)],
        task_id=f"invoice-{invoice.id}",
    )

    log.info("invoice_uploaded", invoice_id=str(invoice.id), filename=file.filename)
    return InvoiceUploadResponse(id=invoice.id, original_filename=invoice.original_filename, status=invoice.status)


@router.get("/stats", response_model=DashboardStats)
def dashboard_stats(db: DBSession, _user: CurrentUser) -> DashboardStats:
    return InvoiceService(db).dashboard_stats()


@router.get("", response_model=Page[InvoiceRead])
def list_invoices(
    db: DBSession,
    _user: CurrentUser,
    status_filter: Optional[InvoiceStatus] = Query(None, alias="status"),
    vendor_name: Optional[str] = None,
    invoice_number: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=200),
    sort_by: str = Query("created_at"),
    sort_dir: str = Query("desc", pattern="^(asc|desc)$"),
) -> Page[InvoiceRead]:
    filters = InvoiceFilters(
        status=status_filter, vendor_name=vendor_name, invoice_number=invoice_number,
        date_from=date_from, date_to=date_to, search=search,
    )
    items, total = InvoiceService(db).list(filters, page=page, size=size, sort_by=sort_by, sort_dir=sort_dir)
    return Page[InvoiceRead](
        items=[InvoiceRead.model_validate(i) for i in items],
        total=total, page=page, size=size,
        pages=max(1, ceil(total / size)) if total else 1,
    )


@router.get("/{invoice_id}", response_model=InvoiceDetail)
def get_invoice(invoice_id: UUID, db: DBSession, _user: CurrentUser) -> InvoiceDetail:
    invoice = InvoiceService(db).get_with_logs(invoice_id)
    return InvoiceDetail.model_validate(invoice)


# ==========================================
# NEW EXCEL EXPORT ENDPOINT
# ==========================================
@router.get("/export/excel", summary="Export Invoices to styled Excel file")
def export_invoices_excel(
    db: DBSession,
    _user: CurrentUser,
    status_filter: Optional[InvoiceStatus] = Query(None, alias="status"),
    vendor_name: Optional[str] = None,
    invoice_number: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    search: Optional[str] = None,
):
    # 1. Fetch filtered data (limit to 10,000 rows for performance)
    filters = InvoiceFilters(
        status=status_filter, vendor_name=vendor_name, invoice_number=invoice_number,
        date_from=date_from, date_to=date_to, search=search,
    )
    items, _ = InvoiceService(db).list(filters, page=1, size=10000)

    # 2. Setup Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processed Invoices"

    # 3. Write and Style Headers
    headers = ["Invoice #", "Vendor", "Date", "Currency", "Subtotal", "Tax", "Total", "Status", "Confidence"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2B579A")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 4. Write Data
    for inv in items:
        ws.append([
            inv.invoice_number or "N/A",
            inv.vendor_name or "N/A",
            str(inv.invoice_date) if inv.invoice_date else "N/A",
            inv.currency or "USD",
            float(inv.subtotal) if inv.subtotal else 0.0,
            float(inv.tax_amount) if inv.tax_amount else 0.0,
            float(inv.total_amount) if inv.total_amount else 0.0,
            inv.status.value if hasattr(inv.status, 'value') else str(inv.status),
            float(inv.confidence_score) if inv.confidence_score else 0.0
        ])

    # 5. Format Data Rows (Colors, Currency, Percentages)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=9):
        row[4].number_format = '"$"#,##0.00'
        row[5].number_format = '"$"#,##0.00'
        row[6].number_format = '"$"#,##0.00'
        row[8].number_format = '0.0%'

        status_val = str(row[7].value).upper()
        if status_val in ["APPROVED", "POSTED"]:
            row[7].fill = PatternFill("solid", fgColor="C6EFCE")
            row[7].font = Font(color="006100")
        elif status_val in ["REVIEW_REQUIRED", "PROCESSING"]:
            row[7].fill = PatternFill("solid", fgColor="FFEB9C")
            row[7].font = Font(color="9C5700")
        elif status_val in ["FAILED", "REJECTED"]:
            row[7].fill = PatternFill("solid", fgColor="FFC7CE")
            row[7].font = Font(color="9C0006")

        for cell in row:
            cell.border = thin_border
            if cell.column in [3, 4, 8, 9]:
                cell.alignment = center_align

    # 6. Auto-adjust columns & Freeze panes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 3

    ws.freeze_panes = "A2"

    # 7. Return via Stream
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Invoice_Export.xlsx"}
    )